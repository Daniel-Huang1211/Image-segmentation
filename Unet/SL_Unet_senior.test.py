# -*- coding: utf-8 -*-
import pathlib
from argparse import ArgumentParser

from matplotlib import pyplot as plt
from tqdm import tqdm
from datetime import datetime
from pathlib import Path
from dataset import *
from RandAugment import *
from torch.utils.data import DataLoader
from utils import createTrainHistory, saveDict, loadDict, get_SGD, predicted_color_result, make_grid
from DiceLoss import BinaryDiceLoss
from GDiceLoss import GeneralizedDiceLoss
from metric import get_F1, get_precision, get_sensitivity, get_specificity
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import segmentation_models_pytorch as smp
import torch
from torch import optim
from torchvision.utils import save_image
import torch.nn.functional as F
import os
import numpy as np
import random
import ssl
import openpyxl
import csv
import matplotlib.pyplot as plt
# from kornia.losses import focal_loss
from focalloss import focal_loss
import pandas as pd
from unet_model_semi import UNet

# 這段程式碼設置了 Python 中的 SSL 上下文，將默認的 HTTPS 上下文更改為未驗證的上下文。
ssl._create_default_https_context = ssl._create_unverified_context  # 可讓其他裝置與伺服器連接

"""這段程式碼用於設置隨機種子，以確保在訓練神經網絡時獲得可重現的結果"""
seed = 1234  # 將 Python、NumPy(是一個 Python 庫，用於處理數組和矩陣等多維數據) 和 PyTorch 的隨機種子設置為相同的值，從而確保每次運行代碼時生成的隨機數序列都是相同的。
torch.backends.cudnn.benchmark = True  # 啟用這一選項可以讓 PyTorch 在運行時優化深度卷積網絡的計算，以提高性能。
torch.backends.cudnn.enabled = True  # 啟用 cuDNN 提供的 CUDA 加速。
torch.backends.cudnn.deterministic = True  # 將 cuDNN 的行為設置為可重現的模式，確保每次運行時使用相同的算法和設置，從而獲得可預測的結果。
torch.manual_seed(seed)  # 設置 PyTorch 的隨機種子，從而確保 PyTorch 中的隨機操作具有相同的起始點。
torch.cuda.manual_seed(seed)  # 設置 CUDA 的隨機種子，確保在使用 CUDA 加速時生成的隨機數序列與 PyTorch 中的一致。
torch.cuda.manual_seed_all(seed)  # 將所有可用的 CUDA 設備的隨機種子設置為相同的值，從而確保在多個 GPU 上運行代碼時獲得相同的結果。
np.random.seed(seed)  # 設置 NumPy 的隨機種子，確保 NumPy 中的隨機數生成器產生的數字序列與 PyTorch 中的一致。
random.seed(seed)  # 設置 Python 的隨機種子，確保 Python 中的隨機操作（如列表的隨機排序）具有相同的起始點。


class SupervisedSegmentation:
    def __init__(self,  # 將以下條件初始化
                 model,
                 device,
                 learning_rate,
                 img_dir,
                 label_dir,
                 valid_img_dir,
                 valid_label_dir,
                 save_dir,
                 l2_decay,
                 batch_size,
                 epochs,
                 cpu_core,
                 train_txt_path,
                 valid_txt_path,
                 metric_item,
                 optim_name,
                 loss,
                 model_path):

        """檢查CUDA是否可用，並設置適當的設備（GPU或CPU）"""
        if torch.cuda.is_available():
            device = torch.device("cuda:{}".format(device))
            torch.cuda.set_device(device)
            # device = torch.device("cuda")
        else:
            device = torch.device("cpu")
        self.device = device
        print(self.device)  # 打印所選擇的設備（GPU或CPU）

        """創建UNet模型的實例並將其移動到所選擇的設備上"""
        seg_model = UNet(1, 1)
        self.model = seg_model.to(device)
        self.model_path = model_path
        self.loss = loss.to(device)
        self.dice_loss = BinaryDiceLoss().to(self.device)

        # Count model's parameters
        print("=" * 50)
        print("student model parameters: {:.2f}M".format(sum(p.numel() for p in self.model.parameters()) / 1e6))
        print("teacher model parameters: {:.2f}M".format(sum(p.numel() for p in self.model.parameters()) / 1e6))
        print("=" * 50)

        # optim = getattr(torch.optim, optim_name)
        # self.optim = optim(self.model.parameters(), lr=learning_rate, weight_decay=l2_decay)

        """ 設置優化器（optimizer）（使用自定義的SGD優化器）"""
        self.optim = get_SGD(net=self.model,
                             lr=learning_rate,  # 學習率
                             weight_decay=l2_decay)  # L2 正則化的權重衰減參數

        # Choose the optimizer and scheduler
        # self.optim = optim.get_SGD(self.model.parameters(), lr=learning_rate)
        # self.scheduler = optim.lr_scheduler.StepLR(self.optim, epochs // 3, gamma=0.1, verbose=True)


        """設置模型保存的目錄，如果保存目錄不存在，則創建它"""
        self.save_dir = save_dir
        if not os.path.isdir(str(self.save_dir)):
            os.makedirs(str(self.save_dir))

        """設置訓練的總epoch數"""
        self.epochs = epochs

        """創建訓練數據集實例，使用給定的文件路徑、圖像目錄、標籤目錄和轉換"""
        self.train_dataset = ImageDataSet_supervised_aug(train_txt_path,
                                                         img_dir,
                                                         label_dir,
                                                         transform=RandAugment_best_2aug_with_ori_img(1, 20))

        """創建驗證數據集實例，使用給定的文件路徑、圖像目錄、標籤目錄"""
        self.valid_dataset = ImageDataSet(valid_txt_path,
                                          valid_img_dir,
                                          valid_label_dir,
                                          sort=True)

        """創建訓練數據加載器"""
        self.train_loader = DataLoader(self.train_dataset,
                                       batch_size=batch_size,
                                       shuffle=True,
                                       num_workers=cpu_core,
                                       pin_memory=True,
                                       drop_last=True)

        """創建驗證數據加載器"""
        self.valid_loader = DataLoader(self.valid_dataset,
                                       batch_size=1,
                                       shuffle=False,
                                       num_workers=cpu_core,
                                       pin_memory=True)

        """創建訓練歷史記錄"""
        self.train_history = createTrainHistory(metric_item)

        """初始化最佳目標和最佳目標的epoch"""
        self.best_target = 0.0
        self.best_target_epoch = 0

    """定義訓練方法"""
    """ 
    1.首先，將模型設置為訓練模式，這將啟用模型中與訓練相關的功能，例如 dropout 和 batch normalization 的更新。
    2.初始化用於計算訓練過程中損失和指標的變數，包括總損失 t_loss，敏感度 t_SE，特異度 t_SP，精確度 t_PR 和 F1 分數 t_F1。
    3.使用 tqdm 創建一個進度條，這將用於顯示訓練過程中的進度。
    4.進入訓練迴圈，對訓練數據進行迭代。self.train_loader 是一個 DataLoader，用於將訓練數據分成批次進行訓練。
    5.在每個迭代中，從 self.train_loader 中獲取一個批次的圖像數據 imgs 和對應的標籤 labels。
    6.將圖像數據和標籤數據移動到指定的設備上，例如 GPU 或 CPU。
    7.通過模型 self.model 執行前向傳播，得到模型的預測結果 pred。
    8.計算損失函數，這裡使用了二元交叉熵損失函數以及定制的 Dice 損失函數。這兩個損失函數的結果相加後作為最終的損失。
    9.計算模型預測結果的特異度（Specificity）、敏感度（Sensitivity）、精確度（Precision）和 F1 分數。
    10.使用 self.optim.zero_grad() 將優化器中所有參數的梯度清零，這是因為 PyTorch 默認會累積梯度，所以在每個迭代步驟之前需要清除之前的梯度。
    11.如果啟用了混合精度訓練，則使用 Apex 库的 amp.scale_loss 函數對損失進行放大處理。
    12.使用反向傳播算法計算損失相對於模型參數的梯度。
    13.使用優化器 self.optim 更新模型的參數，這一步是模型訓練的關鍵步驟，它使用梯度下降算法來調整模型參數，從而最小化損失函數。
    14.更新各項指標的累加值，以便計算整個訓練過程中的平均值。
    15.更新進度條，以顯示訓練過程中的進度。
    16.最後，將整個訓練過程中的平均損失、特異度、敏感度、精確度和 F1 分數記錄到 self.train_history 中，以便後續分析和可視化。
    """

    def train(self):
        self.model.train()  # 將模型設置為訓練模式，這會啟用訓練相關的功能，如 dropout 和 batch normalization 的更新。
        t_loss = 0.0
        t_SE = 0.0
        t_SP = 0.0
        t_PR = 0.0
        t_F1 = 0.0
        with tqdm(total=len(self.train_dataset), desc="train ", unit="img",
                  bar_format='{l_bar}{bar:50}{r_bar}{bar:-10b}') as pbar:  # 使用 tqdm 創建一個進度條，用於顯示訓練進度
            for imgs, labels in self.train_loader:
                imgs, labels = imgs.to(self.device, dtype=torch.float32), labels.to(self.device, dtype=torch.float32)

                pred = self.model(imgs)

                loss = self.loss(torch.sigmoid(pred), labels)
                # Compute loss
                # loss = focal_loss(pred, labels.squeeze(1), alpha=0.25, gamma=2, reduction='mean').unsqueeze(0)
                # loss += dice_loss(pred, labels, True, k=0.75)
                loss += self.dice_loss(torch.sigmoid(pred), labels, k=0.75)
                # loss = self.loss(pred, labels)
                # _, pred = torch.max(pred, dim=1, keepdim=True)
                # _, labels = torch.max(labels, dim=1, keepdim=True)
                # print(pred.shape)
                SP = get_specificity(pred, labels)
                SE = get_sensitivity(pred, labels)
                PR = get_precision(pred, labels)
                F1 = get_F1(pred, labels)

                self.optim.zero_grad()
                loss.backward()
                self.optim.step()

                t_loss += loss.item()
                t_SE += SE
                t_SP += SP
                t_PR += PR
                t_F1 += F1
                pbar.update(imgs.shape[0])
        self.train_history["train"]["loss"].append(t_loss / len(self.train_loader))
        self.train_history["train"]["SE"].append(t_SE / len(self.train_loader))
        self.train_history["train"]["SP"].append(t_SP / len(self.train_loader))
        self.train_history["train"]["PR"].append(t_PR / len(self.train_loader))
        self.train_history["train"]["F1"].append(t_F1 / len(self.train_loader))

    """定義驗證方法"""
    """
    1.`self.model.eval()`: 這行程式碼將模型設置為評估模式。在評估模式下，模型不會計算梯度，並且某些層（例如，Batch Normalization 和 Dropout）的行為會與訓練時有所不同。
    2.`torch.no_grad()`: 使用這個上下文管理器可以確保在驗證過程中不會計算梯度。這有助於提高計算效率，因為在驗證過程中我們不需要更新模型參數。
    3.遍歷驗證數據集：通過迭代驗證集的每個圖像和相應的標籤來進行預測和評估。
    4.計算損失：將模型預測的結果與真實的標籤進行比較，並計算模型在驗證集上的損失。在這個例子中，使用了 `self.loss` 定義的損失函數來計算損失值。此外，可能還包括其他損失項，如 Dice 損失。
    5.計算評估指標：除了損失之外，還計算了模型預測結果的特定評估指標，包括特異性（Specificity）、敏感性（Sensitivity）、精確度（Precision）和 F1 分數。這些指標用於評估模型的性能和效果。
    6.累加損失和評估指標：在整個驗證過程中，將每個圖像的損失和評估指標的值進行累加，以便計算整個驗證集上的平均損失和評估指標。
    7.更新進度條：使用進度條來顯示驗證過程的進度，讓使用者可以實時查看模型在驗證集上的評估進度。
    8.記錄結果：將驗證過程中計算得到的損失和評估指標的平均值記錄到訓練歷史記錄中，以便後續分析和可視化。
    """

    def valid(self):
        self.model.eval()
        v_loss = 0.0
        v_SE = 0.0
        v_SP = 0.0
        v_PR = 0.0
        v_F1 = 0.0
        with torch.no_grad():
            with tqdm(total=len(self.valid_dataset), desc="valid ", unit="img",
                      bar_format='{l_bar}{bar:50}{r_bar}{bar:-10b}') as pbar:
                for imgs, labels in self.valid_loader:
                    imgs, labels = imgs.to(self.device, dtype=torch.float32), labels.to(self.device,
                                                                                        dtype=torch.float32)

                    pred = self.model(imgs)
                    loss = self.loss(torch.sigmoid(pred), labels)
                    loss += self.dice_loss(torch.sigmoid(pred), labels, k=0.75)
                    # loss = self.loss(pred, labels)
                    # _, pred = torch.max(pred, dim=1, keepdim=True)
                    SP = get_specificity(pred, labels)
                    SE = get_sensitivity(pred, labels)
                    PR = get_precision(pred, labels)
                    F1 = get_F1(pred, labels)

                    v_loss += loss.item()
                    v_SE += SE
                    v_SP += SP
                    v_PR += PR
                    v_F1 += F1
                    pbar.update(imgs.shape[0])
        self.train_history["valid"]["loss"].append(v_loss / len(self.valid_loader))
        self.train_history["valid"]["SE"].append(v_SE / len(self.valid_loader))
        self.train_history["valid"]["SP"].append(v_SP / len(self.valid_loader))
        self.train_history["valid"]["PR"].append(v_PR / len(self.valid_loader))
        self.train_history["valid"]["F1"].append(v_F1 / len(self.valid_loader))

    """定義加載模型並保存CSV的方法"""
    """
    1.checkpoint = torch.load(self.model_path, self.device): 這行程式碼負責從指定的路徑 `self.model_path` 加載預先訓練好的模型。
    2.將模型設置為評估模式，以確保模型在進行預測時不會計算梯度，並且可能會更改一些層的行為。
    3.遍歷驗證數據集：對於每個驗證圖像，計算模型的預測結果以及相應的損失和評估指標（如特異性、敏感性、精確度和 F1 分數）。
    4.將模型的預測結果保存為圖像文件，以便進行後續的視覺化和分析。
    5.將模型的預測結果、損失和評估指標以及文件名保存到 CSV 文件中。
    6.在每次迭代完成後，計算並顯示平均推理時間（`Latency_mean`）。
    7.在整個驗證過程完成後，如果是最後一個折（fold），則計算並保存整個驗證集上的平均損失和評估指標，以及每個折的平均 F1 分數和標準差。同時，將最佳 F1 分數及其所在的折數也保存到 CSV 文件中。
    """

    def load_model_and_save_csv(self):
        checkpoint = torch.load(self.model_path, self.device)
        self.model.load_state_dict(checkpoint)
        # current_fold = str(self.save_dir)[-1]
        current_fold = "2"
        csv_file_path = str(self.save_dir) + "{}.csv".format(str(self.save_dir).split("/")[-2])

        self.model.eval()
        latency_sum = []
        v_loss = 0.0
        v_SE = 0.0
        v_SP = 0.0
        v_PR = 0.0
        v_F1 = 0.0
        with torch.no_grad():
            with tqdm(total=len(self.valid_dataset), desc="valid ", unit="img",
                      bar_format='{l_bar}{bar:50}{r_bar}{bar:-10b}') as pbar:

                if current_fold == "1" or True:
                    f = open(csv_file_path, "w")
                    load_model_writer = csv.writer(f)
                    load_model_writer.writerow(["Fold", "filename", "loss", "SP", "SE", "PR", "F1"])
                else:
                    f = open(csv_file_path, "a")
                    load_model_writer = csv.writer(f)
                file_index = 0
                for imgs, labels in self.valid_loader:
                    imgs, labels = imgs.to(self.device, dtype=torch.float32), labels.to(self.device,
                                                                                        dtype=torch.float32)
                    file_name = self.valid_dataset.fileNames[file_index]
                    start = float(datetime.now().strftime("%S.%f")[:-3])
                    pred = self.model(imgs)
                    end = float(datetime.now().strftime("%S.%f")[:-3])
                    latency_delta = end - start
                    latency_sum.append(latency_delta)
                    loss = self.loss(pred, labels)
                    SP = get_specificity(pred, labels)
                    SE = get_sensitivity(pred, labels)
                    PR = get_precision(pred, labels)
                    F1 = get_F1(pred, labels)
                    pred = predicted_color_result(imgs, pred, labels)
                    save_image(pred, self.save_dir / file_name)
                    make_grid(imgs[0, 0, :, :], labels[0, 0, :, :], pred[0, :, :, :],
                              {"SE": SE, "PR": PR, "F1": F1, "file_name": file_name}, str(self.save_dir))

                    load_model_writer.writerow([str(self.save_dir)[-1], file_name, loss.item(), SP, SE, PR, F1])
                    v_loss += loss.item()
                    v_SE += SE
                    v_SP += SP
                    v_PR += PR
                    v_F1 += F1

                    file_index += 1

                    pbar.update(imgs.shape[0])
                Latency_mean = sum(latency_sum[1:]) / (len(self.valid_loader) - 1)
                print(f"Latency_mean : {Latency_mean}")
                f.close()
                if current_fold == "5" or True:
                    total_5_csv = pd.read_csv(csv_file_path)
                    avg_result = [total_5_csv['loss'].mean(),
                                  total_5_csv['SP'].mean(),
                                  total_5_csv['SE'].mean(),
                                  total_5_csv['PR'].mean(),
                                  total_5_csv['F1'].mean()]
                    std_result = [total_5_csv['loss'].std(),
                                  total_5_csv['SP'].std(),
                                  total_5_csv['SE'].std(),
                                  total_5_csv['PR'].std(),
                                  total_5_csv['F1'].std()]
                    fold_F1_avg = [total_5_csv['F1'].loc[:99].mean(),
                                   total_5_csv['F1'].loc[100:199].mean(),
                                   total_5_csv['F1'].loc[200:299].mean(),
                                   total_5_csv['F1'].loc[300:399].mean(),
                                   total_5_csv['F1'].loc[400:499].mean()]
                    fold_F1_avg = [float(i) for i in fold_F1_avg]
                    fold_F1_std = [total_5_csv['F1'].loc[:99].std(),
                                   total_5_csv['F1'].loc[100:199].std(),
                                   total_5_csv['F1'].loc[200:299].std(),
                                   total_5_csv['F1'].loc[300:399].std(),
                                   total_5_csv['F1'].loc[400:499].std()]
                    fold_F1_std = [float(i) for i in fold_F1_std]
                    with open(csv_file_path, "a") as f:
                        load_model_writer = csv.writer(f)
                        load_model_writer.writerow(["", "", "loss_avg", "SP_avg", "SE_avg", "PR_avg", "F1_avg"])
                        load_model_writer.writerow(["", ""] + [round(i, 5) for i in avg_result])
                        load_model_writer.writerow(["", ""] + [round(i, 4) for i in std_result])
                        load_model_writer.writerow(
                            ["", "", "fold1_F1_avg", "fold2_F1_avg", "fold3_F1_avg", "fold4_F1_avg", "fold5_F1_avg"])
                        load_model_writer.writerow(["", ""] + [round(i, 5) for i in fold_F1_avg])
                        load_model_writer.writerow(["", ""] + [round(i, 4) for i in fold_F1_std])
                        load_model_writer.writerow(["", "", "fold_F1_max", "", "", "", ""])
                        load_model_writer.writerow(
                            ["", ""] + [round(max(fold_F1_avg), 5)] + [fold_F1_avg.index(max(fold_F1_avg)) + 1])
                        load_model_writer.writerow(
                            ["", ""] + [round(fold_F1_std[fold_F1_avg.index(max(fold_F1_avg))], 4)])

    """定義計算所有閾值的方法"""
    """
    1.	模型載入：使用 `torch.load` 方法載入預先訓練好的模型的權重。
    2.	初始化：初始化 `current_fold` 變數為 `"2"`，並初始化一個 CSV 文件的路徑 `csv_PR_curve`。
    3.	模型評估：將模型設置為評估模式，並關閉梯度計算以節省資源。
    4.	PR 曲線計算：對於每個閾值（0.1 到 0.9），在驗證數據集上遍歷，計算模型的預測結果，並計算該閾值下的損失、特異性（SP）、敏感性（SE）、精確度（PR）和 F1 值。這些指標用於衡量模型在不同閾值下的性能。
    5.	結果保存：將每個閾值下的 SE 和 PR 存儲到 `pr_curve` 字典中，並使用 `saveDict` 方法保存為 pickle 文件。
    6.	多折計算：如果是最後一個折（fold），則將每個折的 SE 和 PR 加總並計算平均值，最後將這些結果保存為 CSV 文件。
    """

    def all_thresholds(self):
        checkpoint = torch.load(self.model_path, self.device)
        self.model.load_state_dict(checkpoint)
        # current_fold = str(self.save_dir)[-1]
        current_fold = "2"
        csv_PR_curve = str(self.save_dir) + "pr_curve_0.1_0.9.csv"

        self.model.eval()
        with torch.no_grad():
            with tqdm(total=len(self.valid_dataset) * 9, desc="PR Curve ", unit="img",
                      bar_format='{l_bar}{bar:50}{r_bar}{bar:-10b}') as pbar:
                pr_curve = {current_fold: {"SE": [], "PR": []}}  # 放目前的 fold 的 SE 與 PR
                for th in range(1, 10):
                    v_loss = 0.0
                    v_SE = 0.0
                    v_SP = 0.0
                    v_PR = 0.0
                    v_F1 = 0.0
                    file_index = 0
                    th /= 10  # 0.1 ~ 0.9
                    for imgs, labels in self.valid_loader:
                        imgs, labels = imgs.to(self.device, dtype=torch.float32), labels.to(self.device,
                                                                                            dtype=torch.float32)
                        file_name = self.valid_dataset.fileNames[file_index]
                        pred = self.model(imgs)
                        loss = self.loss(torch.sigmoid(pred), labels)
                        loss += self.dice_loss(torch.sigmoid(pred), labels, k=0.75)
                        SP = get_specificity(pred, labels, threshold=th)
                        SE = get_sensitivity(pred, labels, threshold=th)
                        PR = get_precision(pred, labels, threshold=th)
                        F1 = get_F1(pred, labels, threshold=th)

                        v_loss += loss.item()
                        v_SE += SE
                        v_SP += SP
                        v_PR += PR
                        v_F1 += F1
                        file_index += 1

                        pbar.update(imgs.shape[0])
                    pr_curve[current_fold]["SE"].append(v_SE / len(self.valid_loader))  # 新增 SE
                    pr_curve[current_fold]["PR"].append(v_PR / len(self.valid_loader))  # 新增 PR
                # print(pr_curve)
                saveDict("{}/../PR_curve_fold_{}.pickle".format(str(self.save_dir), current_fold), pr_curve)

                if current_fold == "5" or True:
                    # combine pickles
                    FiveFold_SE = np.zeros((1, 9), dtype=np.float32)[0]
                    FiveFold_PR = np.zeros((1, 9), dtype=np.float32)[0]
                    for fold_index in range(2, 3):
                        fold_index = str(fold_index)
                        pickle_data = loadDict("{}/../PR_curve_fold_{}.pickle".format(str(self.save_dir), fold_index))
                        print(pickle_data)
                        FiveFold_SE += np.array(pickle_data[fold_index]["SE"])
                        FiveFold_PR += np.array(pickle_data[fold_index]["PR"])
                    # FiveFold_SE = FiveFold_SE / 5
                    # FiveFold_PR = FiveFold_PR / 5
                    with open(csv_PR_curve, "w") as f:
                        pickle_writer = csv.writer(f)
                        pickle_writer.writerow(["SE", ] + FiveFold_SE.tolist())
                        pickle_writer.writerow(["PR", ] + FiveFold_PR.tolist())
                        #os.system("rm {}/../PR_curve_fold_*.pickle".format(str(self.save_dir)))

    """定義開始訓練的方法"""
    """
    1.訓練開始前的初始化：打印出程式啟動的時間、PyTorch 版本和訓練所使用的設備（device），並檢查是否存在已經訓練好的模型權重文件。如果存在並且用戶同意覆蓋，則會繼續；否則，程式將終止執行。
    2.執行每個 epoch 的訓練和驗證：使用 for 迴圈對每個 epoch 進行迭代。在每個 epoch 中，先打印出當前 epoch 的數量，然後計算訓練和驗證的結果。在訓練過程中，調用 `train()` 方法來進行模型的訓練，並調用 `valid()` 方法來對模型進行驗證。另外，還可能調用學習率的調整方法（`self.scheduler.step()`）。
    3.顯示訓練結果：使用 `show_epoch_results()` 方法來顯示當前 epoch 的訓練結果，包括損失和性能指標。
    4.保存最佳模型：調用 `save_best_target()` 方法來保存在驗證集上表現最佳的模型。
    5.保存訓練歷史：將訓練歷史保存為 pickle 文件。
    6.訓練結束後的清理和統計：打印出訓練結束的時間、總訓練時間等統計信息。
    """

    def start(self):
        plt.close('all')
        begin_time = datetime.now()
        print("*" * 150)
        print("training start at ", begin_time)
        print("Torch Version : ", torch.__version__)
        print("Device : ", self.device)
        print("*" * 150)

        import matplotlib as mpl
        mpl.rcParams['figure.max_open_warning'] = 100

        Unet_parameter = str(self.save_dir) + "parameter.xlsx"
        if os.path.isfile(Unet_parameter):
            os.remove(Unet_parameter)

        if os.path.isfile(str(self.save_dir / "bestF1.pth")):
            overwrite_file = input("{}/bestF1.pth is exist, overwrite?(Y/n) ".format(str(self.save_dir)))
            if overwrite_file != "Y":
                print("stop program")
                exit()
        for epoch in range(self.epochs):
            print("*" * 150)
            print("Epoch %d/%d \n" % (epoch + 1, self.epochs))
            start_time = datetime.now()
            self.train()
            self.valid()
            # self.scheduler.step()
            end_time = datetime.now()
            epoch_total_time = end_time - start_time
            self.show_epoch_results(epoch, epoch_total_time)
            self.save_best_target(epoch)
            print("*" * 120)
            print("Total end time:", datetime.now() + epoch_total_time * (self.epochs - epoch))
            print("*" * 120)
            saveDict("%s/train_history.pickle" % (str(self.save_dir)), self.train_history)
            print("Epoch %d end" % (epoch + 1))
            self.create_xlsx_file(epoch, epoch_total_time)
        finish_time = datetime.now()
        print("*" * 150)
        print("training end at ", finish_time)
        print("Total Training Time : ", finish_time - begin_time)
        print("*" * 150)

        # 在訓練結束後繪製圖表
        self.plot_metrics_from_xlsx()

    def save_best_target(self, epoch):
        if (self.train_history['valid']['F1'][epoch]) > self.best_target:
            # early_stop_count = 0
            self.best_target_epoch = epoch + 1
            last_best = self.best_target
            self.best_target = self.train_history['valid']['F1'][epoch]
            improve_result = f"F1 improves from {last_best:2.5f} to {self.best_target:2.5f}"
            print(improve_result)
            torch.save(self.model.state_dict(), self.save_dir / "bestF1.pth")
            print("save model to %s" % (str(self.save_dir / "bestF1.pth")))
        else:
            improve_result = f"valid_F1 did not improve from {self.best_target:2.5f} " \
                             f"since Epoch {self.best_target_epoch:d}"
            print(improve_result)

    def show_epoch_results(self, epoch, epoch_total_time):
        result = \
            f"Epoch {epoch + 1} time : {epoch_total_time.seconds} secs, " \
            f"loss : {self.train_history['train']['loss'][epoch]:2.5f}, " \
            f"PR : {self.train_history['train']['PR'][epoch]:2.5f}, " \
            f"SE : {self.train_history['train']['SE'][epoch]:2.5f}, " \
            f"SP : {self.train_history['train']['SP'][epoch]:2.5f}, " \
            f"F1 : {self.train_history['train']['F1'][epoch]:2.5f}, " \
            f"valid loss : {self.train_history['valid']['loss'][epoch]:2.5f}, " \
            f"valid_PR : {self.train_history['valid']['PR'][epoch]:2.5f}, " \
            f"valid_SE : {self.train_history['valid']['SE'][epoch]: 2.5f}, " \
            f"valid_SP : {self.train_history['valid']['SP'][epoch]:2.5f}, " \
            f"valid_F1 : {self.train_history['valid']['F1'][epoch]:2.5f}"
        print(result)
    def create_xlsx_file(self, epoch, epoch_total_time):
        # 设置 Excel 文件的路径
        Unet_parameter = str(self.save_dir) + "parameter.xlsx"

        # 如果文件不存在，则创建并写入标题行
        if not os.path.isfile(Unet_parameter):
            field_names = ["Epoch", "Epoch Time (seconds)", "Train Loss", "Train PR", "Train SE", "Train SP",
                           "Train F1", "Valid Loss", "Valid PR", "Valid SE", "Valid SP", "Valid F1"]
            with pd.ExcelWriter(Unet_parameter, engine='openpyxl') as writer:
                pd.DataFrame(columns=field_names).to_excel(writer, index=False)

        # 打开文件以追加模式写入数据
        workbook = load_workbook(Unet_parameter)
        worksheet = workbook.active

        # 将数据写入 Excel 文件
        df = pd.DataFrame([[
            epoch + 1,
            epoch_total_time.seconds,
            self.train_history['train']['loss'][epoch],
            self.train_history['train']['PR'][epoch],
            self.train_history['train']['SE'][epoch],
            self.train_history['train']['SP'][epoch],
            self.train_history['train']['F1'][epoch],
            self.train_history['valid']['loss'][epoch],
            self.train_history['valid']['PR'][epoch],
            self.train_history['valid']['SE'][epoch],
            self.train_history['valid']['SP'][epoch],
            self.train_history['valid']['F1'][epoch]
        ]])
        for r in dataframe_to_rows(df, index=False, header=False):
            worksheet.append(r)

        # 保存 XLSX 文件
        workbook.save(Unet_parameter)

    def create_xlsx_file(self, epoch, epoch_total_time):
        # 设置 Excel 文件的路径
        Unet_parameter = str(self.save_dir) + "parameter.xlsx"

        # 如果文件不存在，则创建并写入标题行
        if not os.path.isfile(Unet_parameter):
            field_names = ["Epoch", "Epoch Time (seconds)", "Train Loss", "Train PR", "Train SE", "Train SP",
                           "Train F1", "Valid Loss", "Valid PR", "Valid SE", "Valid SP", "Valid F1"]
            with pd.ExcelWriter(Unet_parameter, engine='openpyxl') as writer:
                pd.DataFrame(columns=field_names).to_excel(writer, index=False)

        # 打开文件以追加模式写入数据
        workbook = load_workbook(Unet_parameter)
        worksheet = workbook.active

        # 将数据写入 Excel 文件
        df = pd.DataFrame([[
            epoch + 1,
            epoch_total_time.seconds,
            self.train_history['train']['loss'][epoch],
            self.train_history['train']['PR'][epoch],
            self.train_history['train']['SE'][epoch],
            self.train_history['train']['SP'][epoch],
            self.train_history['train']['F1'][epoch],
            self.train_history['valid']['loss'][epoch],
            self.train_history['valid']['PR'][epoch],
            self.train_history['valid']['SE'][epoch],
            self.train_history['valid']['SP'][epoch],
            self.train_history['valid']['F1'][epoch]
        ]])
        for r in dataframe_to_rows(df, index=False, header=False):
            worksheet.append(r)

        # 保存 XLSX 文件
        workbook.save(Unet_parameter)

    def create_xlsx_file(self, epoch, epoch_total_time):
        # 设置 Excel 文件的路径
        Unet_parameter = str(self.save_dir) + "parameter.xlsx"

        # 如果文件不存在，则创建并写入标题行
        if not os.path.isfile(Unet_parameter):
            field_names = ["Epoch", "Epoch Time (seconds)", "Train Loss", "Train PR", "Train SE", "Train SP",
                           "Train F1", "Valid Loss", "Valid PR", "Valid SE", "Valid SP", "Valid F1"]
            with pd.ExcelWriter(Unet_parameter, engine='openpyxl') as writer:
                pd.DataFrame(columns=field_names).to_excel(writer, index=False)

        # 打开文件以追加模式写入数据
        workbook = load_workbook(Unet_parameter)
        worksheet = workbook.active

        # 将数据写入 Excel 文件
        df = pd.DataFrame([[
            epoch + 1,
            epoch_total_time.seconds,
            self.train_history['train']['loss'][epoch],
            self.train_history['train']['PR'][epoch],
            self.train_history['train']['SE'][epoch],
            self.train_history['train']['SP'][epoch],
            self.train_history['train']['F1'][epoch],
            self.train_history['valid']['loss'][epoch],
            self.train_history['valid']['PR'][epoch],
            self.train_history['valid']['SE'][epoch],
            self.train_history['valid']['SP'][epoch],
            self.train_history['valid']['F1'][epoch]
        ]])
        for r in dataframe_to_rows(df, index=False, header=False):
            worksheet.append(r)

        # 保存 XLSX 文件
        workbook.save(Unet_parameter)

    def plot_metrics_from_xlsx(self):
        # 加载 Excel 文件
        Unet_parameter = str(self.save_dir) + "parameter.xlsx"
        df = pd.read_excel(Unet_parameter, engine='openpyxl')

        # 提取所需的列作为 y 轴数据
        epochs = df["Epoch"]
        train_loss = df["Train Loss"]
        train_f1 = df["Train F1"]

        # 创建两个新的图形
        fig1, ax1 = plt.subplots(figsize=(5, 5))
        fig2, ax2 = plt.subplots(figsize=(5, 5))

        # 在第一个图形中绘制 train_loss
        ax1.plot(epochs, train_loss, label="Train Loss")
        ax1.set_title("Training Loss Over Epochs")
        ax1.set_xlabel("Epoch")
        ax1.set_ylabel("Loss")
        ax1.legend()

        # 在第二个图形中绘制 train_f1
        ax2.plot(epochs, train_f1, label="Train F1")
        ax2.set_title("Training F1 Over Epochs")
        ax2.set_xlabel("Epoch")
        ax2.set_ylabel("F1")
        ax2.legend()

        # 将图形保存为临时文件
        fig1.savefig("train_loss_plot.png")
        fig2.savefig("train_f1_plot.png")

        # 将图像插入到 Excel 文件中
        wb = load_workbook(Unet_parameter)
        ws = wb.active
        img1 = Image("train_loss_plot.png")
        img2 = Image("train_f1_plot.png")

        # 插入图像到最后一行数据的下方
        last_row = ws.max_row + 3
        ws.add_image(img1, f"A{last_row}")
        ws.add_image(img2, f"I{last_row}")

        # 调整图像大小和 X 轴间隔
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['I'].width = 20

        # 保存 Excel 文件
        wb.save(Unet_parameter)

"""確保腳本作為主程序運行"""
if __name__ == "__main__":
    """創建解析器對象"""
    """
    1. --b 或 --batch_size：設置批量大小，預設為 10。
    2. --e 或 --Epoch：設置訓練的 epoch 數量，預設為 200。
    3. --c 或 --cpu_core：設置 CPU 核心數，預設為 15。
    4. --d 或 --device：設置訓練所使用的設備，預設為 "3"。該參數可以是 GPU 的編號，例如 "cuda:2"，或者 "cpu"。
    5. --loss 或 --su_loss：設置損失函數，預設為 GeneralizedDiceLoss。
    6. --i_dir 或 --img_dir：設置訓練圖像的目錄路徑。
    7. --l_dir 或 --label_dir：設置訓練標籤的目錄路徑。
    8. --s_dir 或 --save_dir：設置模型和訓練歷史的保存目錄路徑。
    9. --v_i_dir 或 --valid_img_dir：設置驗證圖像的目錄路徑。
    10. --v_l_dir 或 --valid_label_dir：設置驗證標籤的目錄路徑。
    11. --t_txt_path 或 --train_txt_path：設置訓練文本文件的路徑。
    12. --v_txt_path 或 --valid_txt_path：設置驗證文本文件的路徑。
    13. --m 或 --model：設置模型，預設為 UNet(1, 1)。
    14. --lr 或 --learning_rate：設置學習率，預設為 1e-1。
    15. --l2 或 --weight_decay：設置 L2 正則化的權重衰減參數，預設為 1e-4。
    16. --k 或 --metric：設置性能指標，預設為 ["loss", "SE", "SP", "PR", "F1"]。
    17. --model_path：設置模型的保存路徑。
    """
    parser = ArgumentParser()
    parser.add_argument("--b", "--batch_size", default=10, type=int)
    parser.add_argument("--e", "--Epoch", default=10, type=int)
    parser.add_argument("--c", "--cpu_core", default=15)
    parser.add_argument("--d", "--device", default="3")
    # parser.add_argument("--d", "--device",
    #                     default=torch.device("cuda:2" if torch.cuda.is_available() else "cpu"))
    parser.add_argument("--loss", "--su_loss", default=GeneralizedDiceLoss())
    parser.add_argument("--i_dir", "--img_dir",
                        default="/home/jannawu/Desktop/all_cag_files/400/single/4aug/train/5F_2/imgs",
                        type=pathlib.Path)
    parser.add_argument("--l_dir", "--label_dir",
                        default="/home/jannawu/Desktop/all_cag_files/400/single/4aug/train/5F_2/labels",
                        type=pathlib.Path)
    parser.add_argument("--s_dir", "--save_dir",
                        default="./record/Unet/5F_2/",
                        type=pathlib.Path)
    parser.add_argument("--v_i_dir", "--valid_img_dir",
                        default="/home/jannawu/Desktop/all_cag_files/400/single/4aug/valid/5F_2/imgs",
                        type=pathlib.Path)
    parser.add_argument("--v_l_dir", "--valid_label_dir",
                        default="/home/jannawu/Desktop/all_cag_files/400/single/4aug/valid/5F_2/labels",
                        type=pathlib.Path)
    parser.add_argument("--t_txt_path", "--train_txt_path",
                        default="/home/p_01/Desktop/cag_supervised/labeled_400_2.txt")
    parser.add_argument("--v_txt_path", "--valid_txt_path",
                        default="/home/jannawu/Desktop/all_cag_files/400/single/4aug/valid/txt/valid_2.txt")
    parser.add_argument("--m", "--model", default=UNet(1, 1))
    parser.add_argument("--lr", "--learning_rate", default=1e-1, type=float)
    parser.add_argument("--l2", "--weight_decay", default=1e-4)
    parser.add_argument("--k", "--metric", default=["loss", "SE", "SP", "PR", "F1"])
    parser.add_argument("--model_path", default="./record/Unet/5F_2/bestF1.pth")
    args = parser.parse_args()

    """創建SupervisedSegmentation的實例"""
    train = SupervisedSegmentation(model=args.m,
                                   batch_size=args.b,
                                   loss=args.loss,
                                   device=args.d,
                                   learning_rate=args.lr,
                                   l2_decay=args.l2,
                                   img_dir=args.i_dir,
                                   label_dir=args.l_dir,
                                   valid_img_dir=args.v_i_dir,
                                   valid_label_dir=args.v_l_dir,
                                   save_dir=args.s_dir,
                                   cpu_core=args.c,
                                   metric_item=args.k,
                                   epochs=args.e,
                                   train_txt_path=args.t_txt_path,
                                   valid_txt_path=args.v_txt_path,
                                   optim_name="SGD",
                                   model_path=args.model_path,)
    train.start()
    train.load_model_and_save_csv()
    #train.all_thresholds()
